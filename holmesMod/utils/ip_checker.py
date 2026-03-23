import os
import requests
import csv
import socket
import ipaddress
import geoip2.database
import pandas as pd
import glob
import sys
import logging
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side
from termcolor import colored

from .config import get_db_path


# ── Logger setup ─────────────────────────────────────────────────────────────

def setup_logger(log_file_path: str) -> logging.Logger:
    # Create (or retrieve) a file logger that writes to log_file_path
    logger = logging.getLogger("ipcheck")
    logger.setLevel(logging.DEBUG)

    # Avoid adding duplicate handlers when the function is called more than once
    if logger.handlers:
        logger.handlers.clear()

    os.makedirs(os.path.dirname(log_file_path), exist_ok=True)

    fh = logging.FileHandler(log_file_path, encoding="utf-8")
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(
        logging.Formatter("%(asctime)s [%(levelname)s] %(message)s", datefmt="%Y-%m-%d %H:%M:%S")
    )
    logger.addHandler(fh)
    return logger


def _get_log_path(output_file_path: str) -> str:
    # Derive a log file path from the main output file path
    base = os.path.splitext(output_file_path)[0]
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{base}_errors_{timestamp}.log"


# A module-level fallback logger (no-op until setup_logger is called)
logger = logging.getLogger("ipcheck")


# ── Helpers ───────────────────────────────────────────────────────────────────

def outsrc_check(ip_domain):
    try:
        db_path = os.path.join(os.path.dirname(get_db_path('city')), 'outsource_db')

        if not os.path.exists(db_path):
            msg = f"Outsource database directory not found at {db_path}"
            colored_print(f"[!] {msg}", "yellow", "bold")
            logger.warning(msg)
            return "N/A"

        outsrc_files = glob.glob(os.path.join(db_path, "*.txt"))

        if not outsrc_files:
            msg = f"No outsource database files found in {db_path}"
            colored_print(f"[!] {msg}", "yellow", "bold")
            logger.warning(msg)
            return "N/A"

        found_categories = []

        for file in outsrc_files:
            try:
                with open(file, 'r', encoding='utf-8', errors='ignore') as f:
                    content = f.read().splitlines()
                    content = [line.strip() for line in content if line.strip()]

                    if ip_domain in content:
                        category = os.path.basename(file).replace('.txt', '').upper()
                        found_categories.append(category)
            except Exception as e:
                msg = f"Error reading outsource database file {file}: {e}"
                colored_print(f"[!] {msg}", "yellow")
                logger.warning(msg)
                continue

        return ", ".join(found_categories) if found_categories else "N/A"

    except Exception as e:
        msg = f"Error in outsrc_check for '{ip_domain}': {e}"
        colored_print(f"[!] {msg}", "red")
        logger.error(msg)
        return "N/A"


def rdns(ip):
    try:
        hostname, _, _ = socket.gethostbyaddr(ip)
        return hostname
    except (socket.herror, socket.gaierror):
        return "N/A"


def get_ssl_registrar(ip):
    certificate = "N/A"
    registrar = "N/A"
    try:
        if ip == "N/A":
            return certificate, registrar

        api_key = os.environ.get('VT_API_KEY')
        if not api_key:
            msg = "VT_API_KEY environment variable not set"
            colored_print(f"[!] {msg}", "red", "bold")
            logger.error(msg)
            return certificate, registrar

        try:
            ipaddress.ip_address(ip)
            ini_ip = True
        except ValueError:
            ini_ip = False

        headers = {'x-apikey': api_key}

        if ini_ip:
            url = f"https://www.virustotal.com/api/v3/ip_addresses/{ip}"
        else:
            url = f"https://www.virustotal.com/api/v3/domains/{ip}"

        response = requests.get(url, headers=headers)

        if response.status_code == 200:
            data = response.json()
            if 'data' in data and 'attributes' in data['data']:
                attributes = data['data']['attributes']
                if 'last_https_certificate' in attributes:
                    cert_data = attributes['last_https_certificate']

                    if 'subject' in cert_data and 'CN' in cert_data['subject']:
                        certificate = cert_data['subject']['CN']
                    elif 'issuer' in cert_data and 'CN' in cert_data['issuer']:
                        certificate = f"(issuer) {cert_data['issuer']['CN']}"

                if ini_ip:
                    if 'as_owner' in attributes:
                        owner = attributes.get('as_owner', 'Unknown')
                        asn = attributes.get('asn', 'Unknown')
                        registrar = f"AS{asn} ({owner})"
                    elif 'network' in attributes:
                        registrar = f"Network: {attributes.get('network', 'Unknown')}"

                    if certificate == "N/A" and 'last_https_certificate' in attributes:
                        cert_data = attributes['last_https_certificate']
                        if 'subject_alternative_name' in cert_data:
                            alt_names = cert_data['subject_alternative_name']
                            if 'DNS' in alt_names and alt_names['DNS']:
                                certificate = alt_names['DNS'][0]
                                colored_print(f"[+] Using alternative domain from certificate: {certificate}", "green")
                else:
                    if 'whois' in attributes:
                        whois_data = attributes['whois']
                        for line in whois_data.split('\n'):
                            line_lower = line.lower()
                            if 'domain registrar url:' in line_lower or 'registrar url:' in line_lower:
                                parts = line.split(':', 1)
                                if len(parts) > 1:
                                    registrar = parts[1].strip()
                                    break

                        if registrar == "N/A":
                            import re
                            urls = re.findall(r'https?://(?:www\.)?[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+/?[^\s]*', whois_data)
                            for url_found in urls:
                                if any(kw in url_found.lower() for kw in ['registrar', 'whois', 'domain', 'iana', 'icann']):
                                    registrar = url_found
                                    colored_print(f"[+] Found likely registrar URL: {registrar}", "green")
                                    break
        else:
            vt_errors = {
                401: "Authentication failed. Check your API key.",
                403: "Access denied. Check your API key permissions.",
                404: f"No data found for {ip}",
                429: "API request quota exceeded.",
            }
            base_msg = f"Error getting data from VirusTotal for '{ip}': HTTP {response.status_code}"
            detail = vt_errors.get(response.status_code, "")
            colored_print(f"[!] {base_msg}", "yellow")
            if detail:
                colored_print(f"[!] {detail}", "red", "bold")
            logger.error(f"{base_msg}. {detail}".strip())

    except Exception as e:
        msg = f"Error in get_ssl_registrar for '{ip}': {e}"
        colored_print(f"[!] {msg}", "red")
        logger.error(msg)

    return certificate, registrar


def get_ip_info(ip, no_rdns=False):
    city_info = None
    country_info = None
    asn_info = None

    citymmdb = get_db_path('city')
    asnmmdb = get_db_path('asn')
    countmmdb = get_db_path('country')

    if not no_rdns:
        rev_dns = rdns(ip)
        if rev_dns == "N/A":
            msg = f"No reverse DNS found for IP: {ip}"
            colored_print(f"[!] {msg}", 'yellow')
            logger.warning(msg)
    else:
        rev_dns = "N/A"

    try:
        with geoip2.database.Reader(citymmdb) as reader:
            try:
                city_info = reader.city(ip)
            except geoip2.errors.AddressNotFoundError:
                msg = f"No city info found for IP: {ip}"
                colored_print(f"[!] {msg}", 'yellow', 'bold')
                logger.warning(msg)
    except FileNotFoundError:
        msg = f"Database file not found: {citymmdb}"
        colored_print(f"[!] Error: {msg}", 'red', 'bold')
        logger.error(msg)

    try:
        with geoip2.database.Reader(asnmmdb) as reader:
            try:
                asn_info = reader.asn(ip)
            except geoip2.errors.AddressNotFoundError:
                msg = f"No ASN info found for IP: {ip}"
                colored_print(f"[!] {msg}", 'yellow', 'bold')
                logger.warning(msg)
    except FileNotFoundError:
        msg = f"Database file not found: {asnmmdb}"
        colored_print(f"[!] Error: {msg}", 'red', 'bold')
        logger.error(msg)

    try:
        with geoip2.database.Reader(countmmdb) as reader:
            try:
                country_info = reader.country(ip)
            except geoip2.errors.AddressNotFoundError:
                msg = f"No country info found for IP: {ip}"
                colored_print(f"[!] {msg}", 'yellow', 'bold')
                logger.warning(msg)
    except FileNotFoundError:
        msg = f"Database file not found: {countmmdb}"
        colored_print(f"[!] Error: {msg}", 'red', 'bold')
        logger.error(msg)

    if city_info and country_info and asn_info:
        network = 'N/A'
        try:
            network = f"{asn_info.ip_address}/{asn_info.prefix_len}"
        except AttributeError:
            pass

        result = [
            ip,
            city_info.city.names.get('en', 'N/A'),
            city_info.location.latitude if city_info.location.latitude else 'N/A',
            city_info.location.longitude if city_info.location.longitude else 'N/A',
            country_info.country.names.get('en', 'N/A'),
            country_info.country.iso_code if country_info.country.iso_code else 'N/A',
            city_info.continent.names.get('en', 'N/A'),
            asn_info.autonomous_system_number if asn_info else 'N/A',
            asn_info.autonomous_system_organization if asn_info else 'N/A',
            network,
        ]

        if not no_rdns:
            result.append(rev_dns)

        return result

    # One or more DB lookups failed — log it as an error entry
    missing = [name for name, info in [("city", city_info), ("country", country_info), ("ASN", asn_info)] if not info]
    logger.error(f"Incomplete GeoIP data for IP '{ip}' — missing: {', '.join(missing)}. Entry skipped.")
    return None


# ── Core processing helpers ───────────────────────────────────────────────────

def _build_header(no_rdns, virtot, user_agents):
    header = [
        'IP Address', 'IP Category', 'City', 'City Latitude', 'City Longitude',
        'Country', 'Country Code', 'Continent', 'ASN Number', 'ASN Organization', 'Network',
    ]
    if not no_rdns:
        header.append('Reverse DNS')
    if virtot:
        header.extend(['Certificate CN', 'Domain Registrar URL'])
    if user_agents is not None:
        header.append('User Agent')
    return header


def _process_single_entry(entry, idx, virtot, user_agents, no_rdns):
    # Resolve one entry (IP or domain) and return the row list, or None on failure.
    # Also returns a list of error strings encountered during processing.
    entry = entry.strip()
    domain = None
    ip_cat = "N/A"
    errors = []

    try:
        ipaddress.ip_address(entry)
        ip = entry
        if not no_rdns:
            rev_dns = rdns(ip)
            if rev_dns != "N/A":
                domain = rev_dns
        else:
            rev_dns = "N/A"
    except ValueError:
        domain = entry
        try:
            ip = socket.gethostbyname(entry)
            rev_dns = rdns(ip) if not no_rdns else "N/A"
        except socket.gaierror:
            ip_cat = outsrc_check(domain)
            msg = f"Cannot resolve domain: '{entry}' (category: {ip_cat})"
            colored_print(f"[!] Cannot resolve domain: {entry}. Skipping.", 'red', 'bold')
            print(f'But the domain is categorized as {ip_cat}')
            logger.error(msg)
            errors.append(msg)
            return None, errors

    # Category lookup
    ip_cat = outsrc_check(ip)
    if ip_cat == "N/A" and domain and domain != "N/A":
        domain_cat = outsrc_check(domain)
        if domain_cat != "N/A":
            ip_cat = domain_cat

    ip_info = get_ip_info(ip, no_rdns)
    if not ip_info:
        msg = f"Could not retrieve GeoIP information for IP: '{ip}'. Entry skipped."
        colored_print(f"[!] Could not retrieve information for IP: {ip}. Skipping.", 'red')
        logger.error(msg)
        errors.append(msg)
        return None, errors

    ip_info.insert(1, ip_cat)

    if virtot:
        cert_cn, registrar = get_ssl_registrar(domain if domain else ip)
        ip_info.extend([cert_cn, registrar])

    if user_agents is not None:
        ip_info.append(user_agents[idx] if idx < len(user_agents) else "N/A")

    return ip_info, errors


# ── Public API ────────────────────────────────────────────────────────────────

def process_ips_only(ip_list, virtot=False, user_agents=None, no_rdns=False):
    # Process IPs and stream results to stdout only (no file output).
    stdout_writer = csv.writer(sys.stdout, lineterminator='\n')

    for i, entry in enumerate(ip_list):
        row, _ = _process_single_entry(entry, i, virtot, user_agents, no_rdns)
        if row is not None:
            stdout_writer.writerow(row)

    colored_print('\n\n[STAGE-1] Processing completed (no files saved)', 'yellow', 'bold')


def ipcheck_mod(ip_list, output_file_path, virtot=False, user_agents=None, no_rdns=False, no_output=False):
    # ── no_output mode: stdout only, but still set up a logger ───────────────
    if no_output:
        log_path = _get_log_path(output_file_path)
        setup_logger(log_path)
        logger.info("Session started (no_output mode) — errors will be logged to this file.")

        header = _build_header(no_rdns, virtot, user_agents)
        stdout_writer = csv.writer(sys.stdout, lineterminator='\n')
        stdout_writer.writerow(header)
        process_ips_only(ip_list, virtot, user_agents, no_rdns)

        colored_print(f'\n[LOG] Error log saved to: {log_path}', 'cyan', 'bold')
        return

    # ── Normal mode: write CSV + XLSX ────────────────────────────────────────
    results_dir = os.path.dirname(output_file_path)
    os.makedirs(results_dir, exist_ok=True)

    # Resolve output file path (avoid overwriting existing files)
    if os.path.exists(output_file_path):
        base_path = os.path.splitext(output_file_path)[0]
        i = 1
        while os.path.exists(f'{base_path}_v{i}.csv'):
            i += 1
        outfp = f'{base_path}_v{i}.csv'
    else:
        outfp = output_file_path

    # Set up the logger next to the output file
    log_path = _get_log_path(outfp)
    setup_logger(log_path)
    logger.info(f"Session started. Output file: {outfp}")

    header = _build_header(no_rdns, virtot, user_agents)
    stdout_writer = csv.writer(sys.stdout, lineterminator='\n')

    with open(outfp, mode='w', newline='') as file:
        writer = csv.writer(file)
        writer.writerow(header)
        stdout_writer.writerow(header)

        total = len(ip_list)
        skipped = 0

        for i, entry in enumerate(ip_list):
            row, errors = _process_single_entry(entry, i, virtot, user_agents, no_rdns)

            if row is None:
                skipped += 1
                continue

            writer.writerow(row)
            stdout_writer.writerow(row)

    # Summary line in the log
    logger.info(f"Processing complete. Total: {total}, Skipped/Errored: {skipped}, Written: {total - skipped}")

    colored_print('\n\n\n[STAGE-1]', 'yellow', 'bold')
    print(f'Result saved to: {outfp}')
    colored_print(f'[LOG] Error log saved to: {log_path}', 'red', 'bold')
    print("\n")
    create_excel_report(outfp)


# ── Excel export ──────────────────────────────────────────────────────────────

def create_excel_report(csv_file):
    df = pd.read_csv(csv_file)
    excel_file = csv_file.replace('.csv', '.xlsx')
    df.to_excel(excel_file, index=False, engine='openpyxl')

    wb = load_workbook(excel_file)
    ws = wb.active

    alignment = Alignment(horizontal='center', vertical='center')
    border = Border(
        left=Side(border_style='thin'),
        right=Side(border_style='thin'),
        top=Side(border_style='thin'),
        bottom=Side(border_style='thin'),
    )

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = alignment
            cell.border = border

    if 'User Agent' in df.columns:
        ua_col_idx = list(df.columns).index('User Agent') + 1
        ws.column_dimensions[chr(65 + ua_col_idx)].width = 60

    if 'IP Category' in df.columns:
        cat_col_idx = list(df.columns).index('IP Category') + 1
        ws.column_dimensions[chr(65 + cat_col_idx)].width = 20

    wb.save(excel_file)
    colored_print("[STAGE-2]", 'magenta', 'bold')
    print(f'Result saved to: {excel_file}')


# ── Utility ───────────────────────────────────────────────────────────────────

def colored_print(message, color, style=None):
    color_map = {
        'light_yellow': 'yellow', 'light_red': 'red', 'light_green': 'green',
        'light_blue': 'blue', 'light_magenta': 'magenta', 'light_cyan': 'cyan',
    }
    mapped_color = color_map.get(color, color)
    print(colored(message, mapped_color, attrs=[style] if style else []))
